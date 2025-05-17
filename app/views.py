from django.shortcuts import render
from .models import Worker, Product, Groups, DailyProduction
from datetime import datetime, timedelta
from django.db.models import Sum, F, DecimalField, ExpressionWrapper
from django.contrib.auth.decorators import login_required

@login_required
def dashboard(request):
    total_workers = Worker.objects.count()
    total_groups = Groups.objects.count()
    total_products = Product.objects.count()
    today = datetime.today()
    
    start_of_week = today - timedelta(days=today.weekday()) 
    end_of_week = start_of_week + timedelta(days=6)  
    
    start_of_month = today.replace(day=1)  
    end_of_month = today.replace(day=28) + timedelta(days=4)  
    weekly_earnings = DailyProduction.objects.filter(
        date__gte=start_of_week
    ).annotate(
        total=ExpressionWrapper(F('quantity') * F('product__unit_price'), output_field=DecimalField())
    ).aggregate(
        weekly_total=Sum('total')
    )['weekly_total'] or 0

    monthly_earnings = DailyProduction.objects.filter(
        date__gte=start_of_month
    ).annotate(
        total=ExpressionWrapper(F('quantity') * F('product__unit_price'), output_field=DecimalField())
    ).aggregate(
        monthly_total=Sum('total')
    )['monthly_total'] or 0

    weekly_total_quantity = DailyProduction.objects.filter(
        date__range=[start_of_week, end_of_week] 
    ).aggregate(
        total_quantity=Sum('quantity')
    )['total_quantity'] or 0

    monthly_total_quantity = DailyProduction.objects.filter(
        date__range=[start_of_month, end_of_month] 
    ).aggregate(
        total_quantity=Sum('quantity')
    )['total_quantity'] or 0

    group_daily_production = DailyProduction.objects.filter(
        date__range=[start_of_week, end_of_week]
    ).annotate(
        amount_due=ExpressionWrapper(F('quantity') * F('product__unit_price'), output_field=FloatField())
    ).values('worker__groups__name', 'product__item', 'date').annotate(
        total_quantity=Sum('quantity'),
        total_amount_due=Sum('amount_due')
    ).order_by('worker__groups__name', 'date', 'product__item')
    
    context = {
        'total_workers': total_workers,
        'total_groups': total_groups,
        'total_products': total_products,
        'weekly_earnings': weekly_earnings, 
        'monthly_earnings': monthly_earnings,
        'weekly_total_quantity': weekly_total_quantity,
        'monthly_total_quantity': monthly_total_quantity,
        'group_daily_production': group_daily_production,
        'start_of_week': start_of_week,
        'end_of_week': end_of_week,
        'start_of_month': start_of_month,

        'end_of_month': end_of_month,
    }

    return render(request, 'prod/dashboard.html', context)



from .forms import WorkerForm  
@login_required

def edit_worker(request, worker_id):
    worker = get_object_or_404(Worker, id=worker_id)
    
    if request.method == 'POST':
        form = WorkerForm(request.POST, request.FILES, instance=worker)
        if form.is_valid():
            form.save()
            return redirect('manage_workers')  
    else:
        form = WorkerForm(instance=worker)
    
    return render(request, 'prod/edit_worker.html', {'form': form, 'worker': worker})



from django.shortcuts import render, get_object_or_404, redirect
from .models import Product
from django.contrib import messages
@login_required

def edit_product(request, id):
    product = get_object_or_404(Product, id=id)
    
    if request.method == 'POST':
        product.item = request.POST.get('item')
        product.unit_price = request.POST.get('unit_price')
        product.save()
        
        messages.success(request, "Product updated successfully!")
        return redirect('manage_products')  

    return render(request, 'prod/edit_product.html', {'product': product})



from .models import Groups
@login_required

def add_worker(request):
    if request.method == 'POST':
        form = WorkerForm(request.POST, request.FILES)
        if form.is_valid():
            worker = form.save(commit=False)
            worker.groups = Groups.objects.first()  
            worker.save()
            messages.success(request, "Worker added successfully!")
            return redirect('manage_workers')
    else:
        form = WorkerForm()

    return render(request, 'prod/add_worker.html', {'form': form})

from django.shortcuts import render, get_object_or_404, redirect
from .models import Worker
from .forms import WorkerForm, ProductForm
from django.contrib import messages
@login_required

def manage_workers(request):
    workers = Worker.objects.all()

    if request.method == 'POST':
        if 'add_worker' in request.POST:
            form = WorkerForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                messages.success(request, "Worker added successfully!")
                return redirect('manage_workers')
        elif 'edit_worker' in request.POST:
            worker_id = request.POST.get('worker_id')
            worker = get_object_or_404(Worker, id=worker_id)
            form = WorkerForm(request.POST, request.FILES, instance=worker)
            if form.is_valid():
                form.save()
                messages.success(request, "Worker updated successfully!")
                return redirect('manage_workers')
        elif 'delete_worker' in request.POST:
            worker_id = request.POST.get('worker_id')
            worker = get_object_or_404(Worker, id=worker_id)
            worker.delete()
            messages.success(request, "Worker deleted successfully!")
            return redirect('manage_workers')
    
    add_form = WorkerForm()

    return render(request, 'prod/manage_workers.html', {
        'workers': workers,
        'add_form': add_form
    })





from django.shortcuts import render, get_object_or_404, redirect
from .models import Product
from .forms import ProductForm
from django.contrib import messages
@login_required

def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Product added successfully!")
            return redirect('manage_products')
    else:
        form = ProductForm()

    return render(request, 'prod/add_product.html', {'form': form})
@login_required

def manage_products(request):
    products = Product.objects.all()

    if request.method == 'POST':
        if 'edit_product' in request.POST:
            product_id = request.POST.get('product_id')
            product = get_object_or_404(Product, id=product_id)
            form = ProductForm(request.POST, instance=product)
            if form.is_valid():
                form.save()
                messages.success(request, "Product updated successfully!")
                return redirect('manage_products')

        elif 'delete_product' in request.POST:
            product_id = request.POST.get('product_id')
            product = get_object_or_404(Product, id=product_id)
            product.delete()
            messages.success(request, "Product deleted successfully!")
            return redirect('manage_products')

    return render(request, 'prod/manage_products.html', {'products': products})

import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict
from datetime import datetime
from .models import DailyProduction

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from .models import DailyProduction
from datetime import datetime

# Shared styles
bold_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def adjust_column_width(ws, fixed_width_cols=None):
    """
    Adjust column widths based on max length in cells,
    but if fixed_width_cols dict provided, uses those widths.
    """
    fixed_width_cols = fixed_width_cols or {}
    for i, col in enumerate(ws.columns, 1):
        column_letter = get_column_letter(i)
        if column_letter in fixed_width_cols:
            ws.column_dimensions[column_letter].width = fixed_width_cols[column_letter]
            continue
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 3
        ws.column_dimensions[column_letter].width = adjusted_width

# ---------- DAILY EXPORT ----------
def export_daily_production(request):
    queryset = DailyProduction.objects.select_related('worker__groups', 'product')

    daily_group_data = defaultdict(lambda: {'leaders': [], 'non_leaders': []})
    for record in queryset:
        key = (record.date, record.worker.groups.name)
        base_earning = round(record.quantity * float(record.product.unit_price), 2)
        worker_info = {
            'name': record.worker.name,
            'is_leader': record.worker.is_leader,
            'base': base_earning,
            'quantity': record.quantity,
            'description': record.description or '',
            'unit_price': float(record.product.unit_price),
        }
        if record.worker.is_leader:
            daily_group_data[key]['leaders'].append(worker_info)
        else:
            daily_group_data[key]['non_leaders'].append(worker_info)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Production"

    ws.merge_cells('A1:I1')
    ws['A1'] = (
        "This sheet contains daily production records. "
        "Base Earnings = Quantity × Unit Price. Bonuses from non-leaders split equally among leaders (not shown)."
    )
    ws['A1'].font = bold_font
    ws['A1'].alignment = center

    headers = [
        'Date', 'Group', 'Name', 'Is Leader',
        'Base Earnings (GH₵)', 'Final Earnings (GH₵)',
        'Quantity Produced', 'School', 'Unit Price (GH₵)'
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for (date, group), data in daily_group_data.items():
        total_non_leader_base = sum(worker['base'] for worker in data['non_leaders'])
        num_leaders = len(data['leaders'])

        bonus_per_leader = round(total_non_leader_base / num_leaders, 2) if num_leaders > 0 else 0

        for worker in data['leaders']:
            base = worker['base']
            final = round(base + bonus_per_leader, 2)
            ws.append([
                date.strftime('%Y-%m-%d'),
                group,
                worker['name'],
                'Yes',
                base,
                final,
                worker['quantity'],
                worker['description'],
                worker['unit_price']
            ])

        # Add non-leaders rows with no bonus
        for worker in data['non_leaders']:
            base = worker['base']
            final = base
            ws.append([
                date.strftime('%Y-%m-%d'),
                group,
                worker['name'],
                'No',
                base,
                final,
                worker['quantity'],
                worker['description'],
                worker['unit_price']
            ])

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center

    adjust_column_width(ws, fixed_width_cols={'A': 15})

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=daily_production.xlsx'
    wb.save(response)
    return response


# ---------- MONTHLY EXPORT ----------
def export_monthly_summary(request):
    queryset = DailyProduction.objects.select_related('worker__groups', 'product')
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"

    # Description
    ws.merge_cells('A1:G1')
    ws['A1'] = (
        "This sheet summarizes monthly earnings. "
        "Final Payment = Base Earnings + Bonus (if leader). "
        "Leaders receive 50% of total non-leader base earnings as bonus. "
        "Non-leaders take 50% of their base as final payment."
    )
    ws['A1'].font = bold_font
    ws['A1'].alignment = center

    headers = [
        'Month', 'Group', 'Name', 'Is Leader',
        'Base Earnings (GH₵)', 'Bonus (GH₵)', 'Final Earnings (GH₵)'
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    monthly_group_data = defaultdict(lambda: defaultdict(list))

    for record in queryset:
        month = record.date.replace(day=1)
        group = record.worker.groups.name
        monthly_group_data[month][group].append({
            'name': record.worker.name,
            'is_leader': record.worker.is_leader,
            'base': round(record.quantity * float(record.product.unit_price), 2)
        })

    for month, groups in monthly_group_data.items():
        for group, workers in groups.items():
            total_non_leader_base = sum(w['base'] for w in workers if not w['is_leader'])
            leader_bonus = round(0.5 * total_non_leader_base, 2)  # 50% bonus pool for leaders
            num_leaders = sum(1 for w in workers if w['is_leader'])
            bonus_per_leader = round(leader_bonus / num_leaders, 2) if num_leaders > 0 else 0

            for worker in workers:
                is_leader = worker['is_leader']
                base = worker['base']
                bonus = bonus_per_leader if is_leader else 0
                final = round(base + bonus, 2) if is_leader else round(base * 0.5, 2)  # non-leaders get 50%

                ws.append([
                    month.strftime('%b-%Y'), group, worker['name'],
                    'Yes' if is_leader else 'No', round(base, 2),
                    round(bonus, 2), round(final, 2)
                ])

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center

    adjust_column_width(ws, fixed_width_cols={'A': 15})

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=monthly_summary.xlsx'
    wb.save(response)
    return response



from django.contrib.auth.decorators import login_required
from django.db.models import F, DecimalField, ExpressionWrapper
from django.shortcuts import render
from collections import defaultdict

@login_required
def all_earnings_view(request):
    workers = Worker.objects.all()
    groups = Groups.objects.all()

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    group_id = request.GET.get('group')
    worker_search = request.GET.get('worker')
    export_type = request.GET.get('export')

    productions = DailyProduction.objects.all()

    if start_date:
        productions = productions.filter(date__gte=start_date)
    if end_date:
        productions = productions.filter(date__lte=end_date)
    if group_id:
        productions = productions.filter(worker__groups__id=group_id)
    if worker_search:
        productions = productions.filter(worker__name__icontains=worker_search)

    productions = productions.annotate(
        total_earned=ExpressionWrapper(
            F('quantity') * F('product__unit_price'),
            output_field=DecimalField()
        )
    )

    grouped_output = {}

    for group in groups:
        group_productions = productions.filter(worker__groups=group)

        month_groups = defaultdict(list)
        for production in group_productions:
            month = production.date.replace(day=1)
            month_groups[month].append(production)

        for month, prod_list in month_groups.items():
            if month not in grouped_output:
                grouped_output[month] = {}

            if group.name not in grouped_output[month]:
                grouped_output[month][group.name] = {
                    'workers': [],
                    'total_group_earnings': 0,
                    'leader_bonus': 0
                }

            worker_earnings = {}

            for prod in prod_list:
                amount = float(prod.total_earned)
                wid = prod.worker.id
                if wid in worker_earnings:
                    worker_earnings[wid]['amount'] += amount
                else:
                    worker_earnings[wid] = {
                        'worker': prod.worker,
                        'amount': amount
                    }

            all_workers = group.workers.all()
            final_workers_data = []

            total_non_leader_base = 0
            leaders = []

            for worker in all_workers:
                base_amount = worker_earnings.get(worker.id, {'amount': 0})['amount']
                if worker.is_leader:
                    leaders.append(worker)
                else:
                    total_non_leader_base += base_amount

            leader_bonus_pool = 0.5 * total_non_leader_base
            bonus_per_leader = leader_bonus_pool / len(leaders) if leaders else 0

            for worker in all_workers:
                base_amount = worker_earnings.get(worker.id, {'amount': 0})['amount']
                if worker.is_leader:
                    final_amount = base_amount + bonus_per_leader
                    bonus = bonus_per_leader
                else:
                    final_amount = base_amount * 0.5
                    bonus = 0

                final_workers_data.append({
                    'worker': worker,
                    'base_amount': base_amount,
                    'final_amount': final_amount,
                    'bonus': bonus
                })

            grouped_output[month][group.name]['workers'] = final_workers_data
            grouped_output[month][group.name]['leader_bonus'] = leader_bonus_pool
            grouped_output[month][group.name]['total_group_earnings'] = sum(w['final_amount'] for w in final_workers_data)

    if export_type == 'pdf':
        return export_pdf(grouped_output)  # Make sure this is implemented to handle grouped_output

    context = {
        'workers': workers,
        'groups': groups,
        'grouped_output': grouped_output,
        'start_date': start_date,
        'end_date': end_date,
        'worker_search': worker_search,
    }
    return render(request, 'prod/earnings.html', context)




def export_pdf(grouped_output):
    template_path = 'prod/earnings_pdf.html'
    context = {'grouped_output': grouped_output}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="earnings.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors with PDF generation <pre>' + html + '</pre>')
    return response



from xhtml2pdf import pisa
from django.template.loader import get_template
def export_pdf(grouped_output):
    if not grouped_output:
        return HttpResponse("No data to export.")

    template_path = 'prod/earnings_pdf.html'
    context = {'grouped_output': grouped_output}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="earnings.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors with PDF generation <pre>' + html + '</pre>')
    return response


def export_csv(grouped_output):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="earnings.csv"'

    writer = csv.writer(response)
    writer.writerow(['Month', 'Group', 'Name', 'Is Leader', 'Base Earnings', 'Final Earnings', 'Bonus'])

    for month, groups in grouped_output.items():
        for group_name, group_info in groups.items():
            for worker in group_info['workers']:
                writer.writerow([
                    month.strftime('%B %Y'),
                    group_name,
                    worker['worker'].name,
                    'Yes' if worker['worker'].is_leader else 'No',
                    "%.2f" % worker['base_amount'],
                    "%.2f" % worker['final_amount'],
                    "%.2f" % worker['bonus'],
                ])

    return response

from django.shortcuts import render, redirect
from django.db.models import Sum, F, FloatField, ExpressionWrapper
from django.db.models.functions import TruncMonth
from django.contrib.auth.decorators import login_required
from datetime import date

from .models import DailyProduction, Worker, Groups, Product

@login_required
def dashboard_view(request):
    today = date.today()

    if request.method == 'POST':
        group_id = request.POST.get('group')
        worker_id = request.POST.get('worker')
        product_id = request.POST.get('product')
        quantity = request.POST.get('quantity')
        description = request.POST.get('description')  # Get description from form

        if group_id and worker_id and product_id and quantity:
            try:
                worker = Worker.objects.get(id=worker_id)
                product = Product.objects.get(id=product_id)
                quantity = int(quantity)

                # Save the new production record
                DailyProduction.objects.create(
                    worker=worker,
                    product=product,
                    quantity=quantity,
                    description=description,  # Save description
                    date=today
                )

                return redirect('dashboard')
            except (Worker.DoesNotExist, Product.DoesNotExist, ValueError):
                pass  # Optionally handle errors or display messages

    # Query for today's production data and annotate with amount_due
    daily_data = DailyProduction.objects.filter(date=today).annotate(
        amount_due=ExpressionWrapper(
            F('quantity') * F('product__unit_price'),
            output_field=FloatField()
        )
    )

    # Monthly group production summary
    monthly_group_summary = DailyProduction.objects.annotate(
        month=TruncMonth('date'),
        group_name=F('worker__groups__name')
    ).values('month', 'group_name').annotate(
        total_quantity=Sum('quantity'),
        total_due=Sum(
            ExpressionWrapper(
                F('quantity') * F('product__unit_price'),
                output_field=FloatField()
            )
        )
    ).order_by('month', 'group_name')

    context = {
        'today': today,
        'workers': Worker.objects.select_related('groups').all(),
        'daily_data': daily_data,
        'monthly_group_summary': monthly_group_summary,
        'groups': Groups.objects.all(),
        'products': Product.objects.all(),
    }

    return render(request, 'prod/work.html', context)

from django.http import JsonResponse
from .models import Worker

def get_workers_by_group(request, group_id):
    workers = Worker.objects.filter(groups__id=group_id).values('id', 'name')
    return JsonResponse(list(workers), safe=False)


from collections import defaultdict
from django.shortcuts import render, get_object_or_404
from .models import Worker, DailyProduction
from django.contrib.auth.decorators import login_required

@login_required
def worker_detail(request, pk):
    worker = get_object_or_404(Worker, pk=pk)

    daily_records = DailyProduction.objects.filter(worker=worker).order_by('-date')

    grouped_records = defaultdict(list)
    for record in daily_records:
        record.amount_due = record.quantity * record.product.unit_price
        grouped_records[record.date].append(record)

    total_earnings = sum(record.amount_due for record in daily_records)

    return render(request, 'prod/worker_detail.html', {
        'worker': worker,
        'grouped_records': dict(grouped_records),
        'total_earnings': total_earnings,
    })


from django.shortcuts import render, get_object_or_404
from .models import Worker, DailyProduction, Product
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from collections import defaultdict
from datetime import datetime
@login_required
def group_product_summary_matrix(request):
    groups = Groups.objects.all()
    products = Product.objects.all()
    table_data = []

    today = now().date()
    start_of_month = today.replace(day=1)
    
    for group in groups:
        product_totals = []
        total_sum = 0

        for product in products:
            total = DailyProduction.objects.filter(
                worker__groups=group,
                product=product,
                date__gte=start_of_month,  
                date__lte=today
            ).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0

            product_totals.append(total)
            total_sum += total

        table_data.append({
            'group': group.name,
            'totals': product_totals,
            'total_sum': total_sum
        })

    return render(request, 'prod/group_product_summary.html', {
        'products': products,
        'table_data': table_data,
    })



from django.shortcuts import render
from .models import DailyProduction

from django.utils.timezone import now
from django.db.models import Sum
from collections import defaultdict
from calendar import month_name
from django.shortcuts import render

def daily_workerproduct_summary(request):
    record = DailyProduction.objects.all()
    groups = Groups.objects.all()
    products = Product.objects.all()
    current_year = now().year
    months = range(1, 13)
    month_labels = [month_name[m] for m in months]

    group_data = []

    for group in groups:
        month_rows = []

        for month in months:
            product_totals = []

            for product in products:
                total = DailyProduction.objects.filter(
                    worker__groups=group,
                    product=product,
                    date__year=current_year,
                    date__month=month
                ).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0

                product_totals.append(total)

            month_rows.append({
                'month': month_name[month],
                'product_totals': product_totals
            })

        group_data.append({
            'group_name': group.name,
            'month_rows': month_rows
        })

    return render(request, 'prod/daily_summary_table.html', {
        'products': products,
        'month_labels': month_labels,
        'group_data': group_data,
            'records':record,

    })




from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, "Logged in successfully!")
            return redirect('dashboard')  
        else:
            messages.error(request, "Invalid username or password.")
    
    return render(request, 'login.html')  
    
def logout_view(request):
    logout(request)
    return redirect('login')

